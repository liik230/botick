"""
pip install aiogram
pip install openpyxl
7587764277:AAHLABcoUkrsMdVneSqmkqVCLVim53LvxJE
"""

import asyncio
import logging
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from io import BytesIO

from aiogram import Bot, Dispatcher, types
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.client.bot import DefaultBotProperties
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.storage.memory import MemoryStorage

# Токен бота
TOKEN = "7587764277:AAHLABcoUkrsMdVneSqmkqVCLVim53LvxJE"

# Директория для хранения файлов
USER_FILES_DIR = "user_files"

# Создаем диспетчер с использованием памяти для хранения состояний
dp = Dispatcher(storage=MemoryStorage())

# Машина состояний для обработки последовательных вопросов
class DataForm(StatesGroup):
    date = State()
    amount = State()
    purchase = State()
    confirm = State()  # Новое состояние для подтверждения

# Клавиатуры для взаимодействия
confirm_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Продолжить"), KeyboardButton(text="Завершить")],
        [KeyboardButton(text="Предпросмотр файла")],
    ],
    resize_keyboard=True,
)

# Инструкция для пользователя
INSTRUCTIONS = """
Привет! Я бот для учета покупок. Вот как мной пользоваться:

1. Введите команду /start, чтобы начать процесс добавления покупок.
2. Следуйте инструкциям:
   - Сначала введите дату покупки.
   - Затем укажите сумму покупки.
   - В конце введите список покупок через запятую.
3. После каждого добавления вы можете:
   - Нажать "Продолжить" для добавления новых покупок.
   - Нажать "Завершить", чтобы получить файл с итоговыми покупками и обнулить данные.
   - Нажать "Предпросмотр файла", чтобы получить текущий файл с покупками, не завершая процесс.
ВНИМАНИЕ!
    бот не находится на сервере и чтобы он 
"""

# Обработчик команды /start
@dp.message(CommandStart())
async def command_start_handler(message: Message, state: FSMContext) -> None:
    # Инициализация истории покупок, если это первый запуск
    await state.update_data(purchases_history=[])
    # Отправка инструкции
    await message.answer(INSTRUCTIONS)
    await message.answer("Введите дату покупки (например, 2024-10-20):")
    await state.set_state(DataForm.date)

# Обработка даты
@dp.message(DataForm.date)
async def process_date(message: Message, state: FSMContext) -> None:
    try:
        # Проверяем корректность даты
        date = datetime.strptime(message.text, "%Y-%m-%d")
        await state.update_data(date=date)
        await message.answer("Введите сумму покупки:")
        await state.set_state(DataForm.amount)
    except ValueError:
        await message.answer("Неверный формат даты. Введите в формате ГГГГ-ММ-ДД.")

# Обработка суммы
@dp.message(DataForm.amount)
async def process_amount(message: Message, state: FSMContext) -> None:
    try:
        amount = float(message.text)
        await state.update_data(amount=amount)
        await message.answer("Введите покупки через запятую (например, хлеб, молоко, сыр):")
        await state.set_state(DataForm.purchase)
    except ValueError:
        await message.answer("Пожалуйста, введите числовое значение для суммы.")

# Обработка покупок
@dp.message(DataForm.purchase)
async def process_purchase(message: Message, state: FSMContext) -> None:
    purchases = message.text.split(", ")

    # Получаем текущие данные
    data = await state.get_data()
    date = data["date"]
    amount = data["amount"]

    # Добавляем текущие данные в историю покупок
    purchase_record = {
        "date": date,
        "amount": amount,
        "purchases": purchases
    }

    # Обновляем историю в состоянии
    purchases_history = data.get("purchases_history", [])
    purchases_history.append(purchase_record)
    await state.update_data(purchases_history=purchases_history)

    # Подтверждение: продолжить или завершить?
    await message.answer("Данные собраны! Хотите продолжить или завершить?", reply_markup=confirm_keyboard)
    await state.set_state(DataForm.confirm)

# Обработка подтверждения от пользователя
@dp.message(DataForm.confirm)
async def process_confirm(message: Message, state: FSMContext) -> None:
    if message.text.lower() == "продолжить":
        # Пользователь хочет продолжить ввод
        await message.answer("Введите дату следующей покупки (например, 2024-10-20):", reply_markup=types.ReplyKeyboardRemove())
        await state.set_state(DataForm.date)
    elif message.text.lower() == "Предпросмотр файла":
        # Пользователь запросил файл
        data = await state.get_data()
        excel_file = generate_excel_file(data["purchases_history"], message.from_user.id)
        await message.answer_document(excel_file, caption="Вот ваш файл с текущими данными.")
    elif message.text.lower() == "завершить":
        # Получаем историю покупок из состояния и сохраняем в Excel
        data = await state.get_data()
        add_data_to_excel(data["purchases_history"], message.from_user.id)

        # Отправляем файл
        excel_file = get_excel_file(message.from_user.id)
        await message.answer_document(excel_file, caption="Вот ваш итоговый Excel файл с покупками.", reply_markup=types.ReplyKeyboardRemove())

        # Обнуляем файл
        reset_excel_file(message.from_user.id)

        # Сбрасываем состояние и историю
        await state.clear()
    else:
        await message.answer("Выберите действие с помощью кнопок: 'Продолжить', 'Файл' или 'Завершить'.")

def add_data_to_excel(purchases_history: list, user_id: int) -> None:
    """
    Функция для добавления данных в существующий Excel файл или создания нового
    для конкретного пользователя
    """
    user_file_path = get_user_file_path(user_id)
    
    if os.path.exists(user_file_path):
        # Если файл существует, открываем его для дополнения
        wb = load_workbook(user_file_path)
        ws = wb.active
    else:
        # Если файла нет, создаем новый файл и лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Покупки"
        # Добавляем заголовки при создании нового файла
        ws.append(["Дата", "Сумма", "Покупки"])
    
    # Добавляем все записи из истории
    for record in purchases_history:
        ws.append([
            record["date"].strftime("%Y-%m-%d"),
            record["amount"],
            ", ".join(record["purchases"])
        ])
    
    # Сохраняем файл
    wb.save(user_file_path)

def reset_excel_file(user_id: int) -> None:
    """
    Функция для обнуления Excel файла (перезапись пустым файлом)
    для конкретного пользователя
    """
    user_file_path = get_user_file_path(user_id)
    
    # Перезаписываем существующий файл пустой таблицей
    wb = Workbook()
    ws = wb.active
    ws.title = "Покупки"
    ws.append(["Дата", "Сумма", "Покупки"])  # Добавляем заголовки
    wb.save(user_file_path)

def get_excel_file(user_id: int) -> types.BufferedInputFile:
    """
    Функция для подготовки Excel файла к отправке для конкретного пользователя
    """
    user_file_path = get_user_file_path(user_id)

    # Читаем файл с диска и подготавливаем к отправке
    with open(user_file_path, "rb") as file:
        return types.BufferedInputFile(file.read(), filename=f"purchases_history_{user_id}.xlsx")

def generate_excel_file(purchases_history: list, user_id: int) -> types.BufferedInputFile:
    """
    Функция для создания Excel файла из текущей истории в памяти и подготовки его к отправке
    для конкретного пользователя
    """
    # Создаем новый Excel файл в памяти
    wb = Workbook()
    ws = wb.active
    ws.title = "Покупки"
    ws.append(["Дата", "Сумма", "Покупки"])  # Заголовки

    # Добавляем все записи
    for record in purchases_history:
        ws.append([
            record["date"].strftime("%Y-%m-%d"),
            record["amount"],
            ", ".join(record["purchases"])
        ])

    # Сохраняем в BytesIO, чтобы не создавать файл на диске
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    return types.BufferedInputFile(excel_io.read(), filename=f"purchases_history_{user_id}.xlsx")

def get_user_file_path(user_id: int) -> str:
    """
    Функция для получения пути к файлу Excel для конкретного пользователя
    """
    if not os.path.exists(USER_FILES_DIR):
        os.makedirs(USER_FILES_DIR)
    
    return os.path.join(USER_FILES_DIR, f"user_{user_id}_purchases.xlsx")

async def main():
    session = AiohttpSession()
    bot = Bot(TOKEN, session=session, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    await dp.start_polling(bot)

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logging.error("Bot stopped!")
